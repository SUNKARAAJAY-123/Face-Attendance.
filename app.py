"""
app.py - Flask Admin Dashboard with Role-Based Authentication
Run:  python app.py
      FLASK_DEBUG=true python app.py

Default credentials:
  Admin   → username: admin    password: admin123
  Teacher → username: teacher  password: teacher123
"""

import os, json, csv, io
from datetime import datetime, timedelta
from functools import wraps

from flask import (Flask, render_template, redirect, url_for,
                   request, flash, jsonify, session, send_file, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import check_password_hash, generate_password_hash

from config       import FLASK_DEBUG, FLASK_PORT, ATTENDANCE_DIR
from database     import init_db, get_db
from models       import (User, get_all_students, get_student,
                           get_attendance_stats, get_student_profile,
                           get_daily_trend, get_unread_notifications,
                           mark_notifications_read, generate_alerts,
                           get_student_by_name)
from data_manager import load_all_attendance, list_persons
from logger_setup import get_logger
from scheduler    import start_scheduler, stop_scheduler

logger = get_logger("app")

# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-me-in-production-32chars!")

login_manager = LoginManager(app)
login_manager.login_view       = "login"
login_manager.login_message    = "Please log in to access the dashboard."
login_manager.login_message_category = "warning"

@login_manager.user_loader
def load_user(user_id):
    return User.get(int(user_id))

# ── Role decorators ────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ── Auth routes ────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user     = User.get_by_username(username)
        conn     = get_db()
        row      = conn.execute(
            "SELECT password_hash FROM users WHERE username=?", (username,)
        ).fetchone()
        conn.close()
        if row and check_password_hash(row["password_hash"], password) and user:
            login_user(user, remember=True)
            logger.info("Login: %s (%s)", username, user.role)
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logger.info("Logout: %s", current_user.username)
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ── Dashboard ──────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    generate_alerts()
    stats   = get_attendance_stats()
    trend   = get_daily_trend(30)
    notifs  = get_unread_notifications()
    return render_template("dashboard.html",
                           stats=stats, trend=trend,
                           notifs=notifs, user=current_user)


# ── Students ───────────────────────────────────────────────────────────────────
@app.route("/students")
@login_required
def students():
    dept   = request.args.get("dept", "")
    search = request.args.get("q", "").lower()
    all_s  = get_all_students()
    if dept:
        all_s = [s for s in all_s if s["department"] == dept]
    if search:
        all_s = [s for s in all_s if search in s["name"].lower()
                                   or search in s["roll_number"].lower()]
    depts  = sorted(set(s["department"] for s in get_all_students()))
    return render_template("students.html", students=all_s,
                           depts=depts, user=current_user)


@app.route("/students/add", methods=["GET", "POST"])
@login_required
def add_student():
    if request.method == "POST":
        name   = request.form["name"].strip()
        roll   = request.form["roll_number"].strip()
        dept   = request.form["department"].strip()
        email  = request.form.get("email", "").strip()
        conn   = get_db()
        try:
            conn.execute(
                "INSERT INTO students(name,roll_number,department,email) VALUES(?,?,?,?)",
                (name, roll, dept, email))
            conn.commit()
            flash(f"Student '{name}' added successfully.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("students"))
    return render_template("add_student.html", user=current_user)


@app.route("/students/<int:sid>")
@login_required
def student_profile(sid):
    profile = get_student_profile(sid)
    if not profile:
        abort(404)
    return render_template("student_profile.html",
                           profile=profile, user=current_user)


@app.route("/students/<int:sid>/delete", methods=["POST"])
@login_required
@admin_required
def delete_student(sid):
    conn = get_db()
    conn.execute("UPDATE students SET is_active=0 WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    flash("Student removed.", "success")
    return redirect(url_for("students"))


# ── Attendance ─────────────────────────────────────────────────────────────────
@app.route("/attendance")
@login_required
def attendance():
    date   = request.args.get("date", datetime.now().strftime("%d-%m-%Y"))
    dept   = request.args.get("dept", "")
    search = request.args.get("q", "").lower()

    conn = get_db()
    rows = conn.execute("""
        SELECT a.*, s.name, s.roll_number, s.department
        FROM attendance a
        JOIN students s ON a.student_id=s.id
        WHERE a.date=?
        ORDER BY a.time_in
    """, (date,)).fetchall()
    conn.close()

    records = [dict(r) for r in rows]
    if dept:
        records = [r for r in records if r["department"] == dept]
    if search:
        records = [r for r in records if search in r["name"].lower()
                                       or search in r["roll_number"].lower()]
    depts = sorted(set(s["department"] for s in get_all_students()))
    return render_template("attendance.html",
                           records=records, date=date,
                           depts=depts, user=current_user)


# ── Analytics ──────────────────────────────────────────────────────────────────
@app.route("/analytics")
@login_required
def analytics():
    stats = get_attendance_stats()
    trend = get_daily_trend(30)
    return render_template("analytics.html",
                           stats=stats, trend=trend, user=current_user)


# ── Low attendance ─────────────────────────────────────────────────────────────
@app.route("/low-attendance")
@login_required
def low_attendance():
    stats = get_attendance_stats()
    return render_template("low_attendance.html",
                           below_75=stats["below_75"], user=current_user)


# ── Notifications ──────────────────────────────────────────────────────────────
@app.route("/notifications")
@login_required
def notifications():
    notifs = get_unread_notifications()
    mark_notifications_read()
    return render_template("notifications.html",
                           notifs=notifs, user=current_user)


# ── User management (admin only) ───────────────────────────────────────────────
@app.route("/scheduler")
@login_required
@admin_required
def scheduler_page():
    import os
    from config import BACKUP_DIR
    backups = []
    if os.path.exists(BACKUP_DIR):
        for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if f.endswith(".zip"):
                fp = os.path.join(BACKUP_DIR, f)
                backups.append({
                    "name":  f,
                    "size":  os.path.getsize(fp) // 1024,
                    "mtime": datetime.fromtimestamp(
                        os.path.getmtime(fp)).strftime("%d-%m-%Y %H:%M"),
                })
    return render_template("scheduler_page.html",
                           backups=backups, user=current_user)


@app.route("/users")
@login_required
@admin_required
def users():
    conn = get_db()
    rows = conn.execute("SELECT id,username,role,full_name,email,is_active FROM users").fetchall()
    conn.close()
    return render_template("users.html",
                           users=[dict(r) for r in rows], user=current_user)


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@admin_required
def add_user():
    if request.method == "POST":
        uname = request.form["username"].strip()
        pwd   = request.form["password"]
        role  = request.form["role"]
        fname = request.form["full_name"].strip()
        email = request.form.get("email", "").strip()
        conn  = get_db()
        try:
            conn.execute(
                "INSERT INTO users(username,password_hash,role,full_name,email) VALUES(?,?,?,?,?)",
                (uname, generate_password_hash(pwd), role, fname, email))
            conn.commit()
            flash(f"User '{uname}' created.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("users"))
    return render_template("add_user.html", user=current_user)


# ── Export routes ──────────────────────────────────────────────────────────────
@app.route("/export/csv")
@login_required
def export_csv():
    records = load_all_attendance()
    si      = io.StringIO()
    writer  = csv.DictWriter(si, fieldnames=["NAME","TIME","DAY","DATE"])
    writer.writeheader()
    writer.writerows(records)
    output = io.BytesIO(si.getvalue().encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv",
                     as_attachment=True,
                     download_name=f"attendance_{datetime.now().strftime('%Y%m%d')}.csv")


@app.route("/export/excel")
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    records = load_all_attendance()
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["#", "Name", "Date", "Day", "Time"]
    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(records, 1):
        ws.append([i, r.get("NAME",""), r.get("DATE",""),
                   r.get("DAY",""),  r.get("TIME","")])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"attendance_{datetime.now().strftime('%Y%m%d')}.xlsx")


@app.route("/export/pdf")
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    records = load_all_attendance()
    output  = io.BytesIO()
    doc     = SimpleDocTemplate(output, pagesize=A4)
    styles  = getSampleStyleSheet()
    elems   = []

    elems.append(Paragraph("Attendance Report", styles["Title"]))
    elems.append(Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elems.append(Spacer(1, 12))

    data = [["#", "Name", "Date", "Day", "Time"]]
    for i, r in enumerate(records, 1):
        data.append([i, r.get("NAME",""), r.get("DATE",""),
                     r.get("DAY",""), r.get("TIME","")])

    t = Table(data, colWidths=[30, 150, 90, 90, 90])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f5f9ff")]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
    ]))
    elems.append(t)
    doc.build(elems)
    output.seek(0)
    return send_file(output, mimetype="application/pdf",
                     as_attachment=True,
                     download_name=f"attendance_{datetime.now().strftime('%Y%m%d')}.pdf")


# ── JSON APIs ──────────────────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def api_stats():
    """Feature 9: polled every 30s by dashboard JS for auto-refresh."""
    return jsonify(get_attendance_stats())


@app.route("/api/insights")
@login_required
def api_insights():
    """Feature 10: return latest AI insight notifications."""
    from database import get_db
    conn  = get_db()
    rows  = conn.execute("""
        SELECT message, created_at FROM notifications
        WHERE type='ai_insight'
        ORDER BY created_at DESC LIMIT 10
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/scheduler/run-now/<job>", methods=["POST"])
@login_required
@admin_required
def run_job_now(job):
    """Manually trigger a scheduler job (admin only)."""
    from scheduler import (job_defaulter_check, job_daily_email_report,
                            job_nightly_backup, job_monthly_report,
                            job_ai_insights)
    jobs = {
        "defaulter": job_defaulter_check,
        "email":     job_daily_email_report,
        "backup":    job_nightly_backup,
        "monthly":   job_monthly_report,
        "insights":  job_ai_insights,
    }
    if job not in jobs:
        return jsonify({"error": "Unknown job"}), 400
    try:
        jobs[job]()
        return jsonify({"status": "ok", "job": job})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    days = int(request.args.get("days", 30))
    return jsonify(get_daily_trend(days))


@app.route("/api/attendance")
@login_required
def api_attendance():
    return jsonify(load_all_attendance())


@app.route("/api/persons")
@login_required
def api_persons():
    return jsonify(list_persons())


@app.route("/health")
def health():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})


# ── Error handlers ─────────────────────────────────────────────────────────────
@app.errorhandler(403)
def forbidden(e):
    return render_template("error.html", code=403,
                           msg="You don't have permission to access this page."), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", code=404,
                           msg="Page not found."), 404


# ── Sync CSV attendance into SQLite on startup ─────────────────────────────────
def sync_csv_to_db():
    """Import any existing CSV attendance records into SQLite."""
    records = load_all_attendance()
    if not records:
        return
    conn = get_db()
    imported = 0
    for r in records:
        name = r.get("NAME","").strip()
        if not name:
            continue
        student = get_student_by_name(name)
        if not student:
            row = conn.execute(
                "SELECT id FROM students WHERE name=?", (name,)
            ).fetchone()
            if not row:
                conn.execute(
                    "INSERT INTO students(name,roll_number,department) VALUES(?,?,?)",
                    (name, name.replace(" ","").upper()[:8], "General"))
                conn.commit()
            student = dict(conn.execute(
                "SELECT * FROM students WHERE name=?", (name,)
            ).fetchone())

        existing = conn.execute(
            "SELECT 1 FROM attendance WHERE student_id=? AND date=?",
            (student["id"], r.get("DATE",""))
        ).fetchone()
        if not existing:
            conn.execute("""
                INSERT INTO attendance(student_id,date,time_in,day_name,status)
                VALUES(?,?,?,?,?)
            """, (student["id"], r.get("DATE",""), r.get("TIME",""),
                  r.get("DAY",""), "Present"))
            imported += 1
    conn.commit()
    conn.close()
    if imported:
        logger.info(f"Synced {imported} CSV records into SQLite.")


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    sync_csv_to_db()
    start_scheduler()
    import atexit
    atexit.register(stop_scheduler)
    logger.info(f"Starting on http://127.0.0.1:{FLASK_PORT}  debug={FLASK_DEBUG}")
    app.run(debug=FLASK_DEBUG, port=FLASK_PORT, use_reloader=False)
