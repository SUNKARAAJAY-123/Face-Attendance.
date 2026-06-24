"""
scheduler.py - APScheduler background jobs

Feature 2:  Daily 6 PM  — auto defaulter detection & notifications
Feature 3:  Daily 7 PM  — email attendance report to admin/teacher
Feature 4:  On-demand   — parent alert when student drops below 75%
Feature 5:  Daily 11 PM — backup DB + face data + attendance CSVs
Feature 7:  1st of month— monthly PDF/Excel report auto-generation
Feature 9:  Every 30s   — dashboard stats cache refresh (via API)
Feature 10: Daily 6 PM  — AI insights (trend analysis, prediction)
"""

import os
import io
import zipfile
import shutil
import smtplib
import logging
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from email.mime.base      import MIMEBase
from email                import encoders

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron         import CronTrigger

from config import (BACKUP_DIR, ATTENDANCE_DIR, DATA_DIR,
                    EMAIL_ENABLED, EMAIL_HOST, EMAIL_PORT,
                    EMAIL_USER, EMAIL_PASS, EMAIL_ADMIN,
                    DEFAULTER_CHECK_TIME, DAILY_REPORT_TIME,
                    BACKUP_TIME, MONTHLY_REPORT_DAY,
                    LOW_ATTENDANCE_PCT)
from logger_setup import get_logger

logger    = get_logger("scheduler")
_scheduler = None


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2 — Auto Defaulter Detection
# ═══════════════════════════════════════════════════════════════════════════════
def job_defaulter_check():
    """Run at 6 PM — recalculate attendance % and create notifications."""
    logger.info("[SCHEDULER] Running defaulter check…")
    try:
        from models import get_attendance_stats, generate_alerts
        generate_alerts()
        stats = get_attendance_stats()
        count = len(stats["below_75"])
        logger.info(f"[SCHEDULER] Defaulter check done — {count} below {LOW_ATTENDANCE_PCT}%")

        # Feature 10 — AI insights at same time
        job_ai_insights(stats)
    except Exception as e:
        logger.error(f"[SCHEDULER] Defaulter check failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 3 — Daily Email Report
# ═══════════════════════════════════════════════════════════════════════════════
def send_email(to: str, subject: str, body_html: str,
               attachments: list = None):
    """Generic email sender using smtplib."""
    if not EMAIL_ENABLED:
        logger.info(f"[EMAIL] Skipped (EMAIL_ENABLED=false). Would send: {subject}")
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_USER
    msg["To"]      = to
    msg.attach(MIMEText(body_html, "html"))

    if attachments:
        for fname, fdata in attachments:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(fdata)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
            msg.attach(part)

    try:
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as s:
            s.starttls()
            s.login(EMAIL_USER, EMAIL_PASS)
            s.sendmail(EMAIL_USER, to, msg.as_string())
        logger.info(f"[EMAIL] Sent '{subject}' to {to}")
    except Exception as e:
        logger.error(f"[EMAIL] Failed: {e}")


def job_daily_email_report():
    """Run at 7 PM — send attendance summary email."""
    logger.info("[SCHEDULER] Generating daily email report…")
    try:
        from models import get_attendance_stats
        stats = get_attendance_stats()
        today = stats["today"]

        below = "".join(
            f"<li>{s['name']} — <strong style='color:#dc2626'>{s['pct']}%</strong></li>"
            for s in stats["below_75"]
        ) or "<li>None ✅</li>"

        body = f"""
        <html><body style='font-family:Segoe UI,sans-serif;color:#1e293b'>
        <h2 style='color:#1a73e8'>📊 Daily Attendance Report — {today}</h2>
        <table style='border-collapse:collapse;width:100%;max-width:500px'>
          <tr><td style='padding:10px;background:#f1f5f9'><b>Total Students</b></td>
              <td style='padding:10px'>{stats['total_students']}</td></tr>
          <tr><td style='padding:10px;background:#dcfce7'><b>Present Today</b></td>
              <td style='padding:10px;color:#16a34a'><b>{stats['present_today']}</b></td></tr>
          <tr><td style='padding:10px;background:#fee2e2'><b>Absent Today</b></td>
              <td style='padding:10px;color:#dc2626'><b>{stats['absent_today']}</b></td></tr>
          <tr><td style='padding:10px;background:#f1f5f9'><b>Avg Attendance</b></td>
              <td style='padding:10px'>{stats['avg_attendance']}%</td></tr>
        </table>
        <h3 style='color:#dc2626;margin-top:20px'>⚠ Students Below 75%</h3>
        <ul>{below}</ul>
        <p style='color:#64748b;font-size:.85rem;margin-top:20px'>
          — AttendanceAI Automated Report</p>
        </body></html>"""

        if EMAIL_ADMIN:
            send_email(EMAIL_ADMIN,
                       f"Attendance Report — {today}", body)
    except Exception as e:
        logger.error(f"[SCHEDULER] Daily email failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 4 — Parent Alert (called on-demand when < 75%)
# ═══════════════════════════════════════════════════════════════════════════════
def send_parent_alert(student_name: str, pct: float, parent_email: str = None):
    """Send low-attendance warning email to parent."""
    if not parent_email:
        logger.info(f"[PARENT ALERT] No email for {student_name} — skipped.")
        return

    body = f"""
    <html><body style='font-family:Segoe UI,sans-serif'>
    <div style='background:#fee2e2;border-left:4px solid #dc2626;
                padding:20px;border-radius:8px;max-width:480px'>
      <h3 style='color:#dc2626'>⚠ Attendance Warning</h3>
      <p>Dear Parent,</p>
      <p>Your ward <strong>{student_name}</strong> has only
         <strong style='color:#dc2626'>{pct}%</strong> attendance.</p>
      <p>The minimum required is <strong>75%</strong>.
         Please ensure regular attendance.</p>
      <p style='color:#64748b;font-size:.85rem'>— AttendanceAI System</p>
    </div></body></html>"""

    send_email(parent_email,
               f"⚠ Low Attendance Warning — {student_name}", body)
    logger.info(f"[PARENT ALERT] Sent to {parent_email} for {student_name}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 5 — Nightly Backup
# ═══════════════════════════════════════════════════════════════════════════════
def job_nightly_backup():
    """Run at 11 PM — zip DB + face data + attendance CSVs."""
    logger.info("[SCHEDULER] Starting nightly backup…")
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts       = datetime.now().strftime("%d_%m_%Y")
        zip_path = os.path.join(BACKUP_DIR, f"backup_{ts}.zip")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # DB
            if os.path.exists("attendance.db"):
                zf.write("attendance.db")

            # Face data
            for root, _, files in os.walk(DATA_DIR):
                for f in files:
                    fp = os.path.join(root, f)
                    zf.write(fp, os.path.relpath(fp))

            # Attendance CSVs
            if os.path.exists(ATTENDANCE_DIR):
                for root, _, files in os.walk(ATTENDANCE_DIR):
                    for f in files:
                        fp = os.path.join(root, f)
                        zf.write(fp, os.path.relpath(fp))

        size_kb = os.path.getsize(zip_path) // 1024
        logger.info(f"[SCHEDULER] Backup saved → {zip_path} ({size_kb} KB)")

        # Keep only last 7 backups
        zips = sorted(
            [f for f in os.listdir(BACKUP_DIR) if f.endswith(".zip")]
        )
        for old in zips[:-7]:
            os.remove(os.path.join(BACKUP_DIR, old))
            logger.info(f"[SCHEDULER] Removed old backup: {old}")

    except Exception as e:
        logger.error(f"[SCHEDULER] Backup failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 7 — Monthly PDF + Excel Report
# ═══════════════════════════════════════════════════════════════════════════════
def job_monthly_report():
    """Run on 1st of each month — generate PDF + Excel and email."""
    logger.info("[SCHEDULER] Generating monthly report…")
    try:
        from models import get_attendance_stats
        from data_manager import load_all_attendance
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from reportlab.lib.pagesizes   import A4
        from reportlab.lib             import colors
        from reportlab.platypus        import (SimpleDocTemplate, Table,
                                               TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles      import getSampleStyleSheet

        stats   = get_attendance_stats()
        records = load_all_attendance()
        month   = datetime.now().strftime("%B %Y")

        # ── Excel ──────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        hfill = PatternFill("solid", fgColor="1a73e8")
        hfont = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(["#","Name","Present","Total","Attendance %","Status"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal="center")
        for i, s in enumerate(sorted(stats["top_10"]+stats["below_75"],
                                     key=lambda x: x["pct"], reverse=True), 1):
            ws.append([i, s["name"], s["present"], s["total"], s["pct"],
                       "Good" if s["pct"]>=75 else "Defaulter"])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        xl_buf = io.BytesIO(); wb.save(xl_buf); xl_bytes = xl_buf.getvalue()

        # ── PDF ────────────────────────────────────────────────────────
        pdf_buf = io.BytesIO()
        doc  = SimpleDocTemplate(pdf_buf, pagesize=A4)
        styl = getSampleStyleSheet()
        elms = [
            Paragraph(f"Monthly Attendance Report — {month}", styl["Title"]),
            Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styl["Normal"]),
            Spacer(1,12),
            Paragraph(f"Total Students: {stats['total_students']}  |  "
                      f"Avg Attendance: {stats['avg_attendance']}%  |  "
                      f"Defaulters: {len(stats['below_75'])}", styl["Normal"]),
            Spacer(1,12),
        ]
        data = [["#","Name","Present","Total","Pct","Status"]]
        for i, s in enumerate(sorted(stats["top_10"]+stats["below_75"],
                                     key=lambda x: x["pct"], reverse=True), 1):
            data.append([i, s["name"], s["present"], s["total"],
                         f"{s['pct']}%", "Good" if s["pct"]>=75 else "Defaulter"])
        t = Table(data, colWidths=[25,130,60,60,50,70])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR", (0,0),(-1,0), colors.white),
            ("FONTNAME",  (0,0),(-1,0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#f5f9ff")]),
            ("GRID",(0,0),(-1,-1),0.5,colors.lightgrey),
            ("FONTSIZE",(0,0),(-1,-1),9),
        ]))
        elms.append(t)
        doc.build(elms)
        pdf_bytes = pdf_buf.getvalue()

        # Save locally
        os.makedirs(BACKUP_DIR, exist_ok=True)
        month_slug = datetime.now().strftime("%Y_%m")
        with open(os.path.join(BACKUP_DIR, f"monthly_{month_slug}.pdf"), "wb") as f:
            f.write(pdf_bytes)
        with open(os.path.join(BACKUP_DIR, f"monthly_{month_slug}.xlsx"), "wb") as f:
            f.write(xl_bytes)

        # Email with attachments
        if EMAIL_ADMIN:
            body = f"<h2>Monthly Attendance Report — {month}</h2>" \
                   f"<p>Please find the PDF and Excel reports attached.</p>" \
                   f"<p>Total students: {stats['total_students']} | " \
                   f"Avg: {stats['avg_attendance']}% | " \
                   f"Defaulters: {len(stats['below_75'])}</p>"
            send_email(EMAIL_ADMIN,
                       f"Monthly Report — {month}", body,
                       attachments=[
                           (f"monthly_{month_slug}.pdf",  pdf_bytes),
                           (f"monthly_{month_slug}.xlsx", xl_bytes),
                       ])

        logger.info(f"[SCHEDULER] Monthly report saved for {month}.")
    except Exception as e:
        logger.error(f"[SCHEDULER] Monthly report failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 10 — Smart AI Insights
# ═══════════════════════════════════════════════════════════════════════════════
def job_ai_insights(stats=None):
    """
    Analyse attendance trends and predict future defaulters.
    Saves insights to DB notifications table.
    """
    logger.info("[SCHEDULER] Generating AI insights…")
    try:
        if stats is None:
            from models import get_attendance_stats
            stats = get_attendance_stats()

        from database import get_db
        conn = get_db()

        insights = []

        # Most punctual student
        if stats["top_10"]:
            top = stats["top_10"][0]
            insights.append(
                f"🌟 Most punctual: {top['name']} with {top['pct']}% attendance."
            )

        # Most absent student
        all_s = stats["top_10"] + stats["below_75"]
        if all_s:
            worst = min(all_s, key=lambda x: x["pct"])
            insights.append(
                f"⚠ Most absent: {worst['name']} with only {worst['pct']}%."
            )

        # Trend: compare last 7 days vs 7 days before
        rows7  = conn.execute("""
            SELECT COUNT(*) as cnt FROM attendance
            WHERE date >= date('now','-7 days') AND status='Present'
        """).fetchone()
        rows14 = conn.execute("""
            SELECT COUNT(*) as cnt FROM attendance
            WHERE date >= date('now','-14 days')
              AND date <  date('now','-7 days')  AND status='Present'
        """).fetchone()
        c7, c14 = (rows7["cnt"] if rows7 else 0), (rows14["cnt"] if rows14 else 0)
        if c14 > 0:
            change = round((c7 - c14) / c14 * 100, 1)
            trend_icon = "📈" if change >= 0 else "📉"
            insights.append(
                f"{trend_icon} Attendance {'improved' if change>=0 else 'dropped'} "
                f"by {abs(change)}% vs last week."
            )

        # Predicted defaulters (students between 75–80% — at risk)
        at_risk = [s for s in all_s if 75 <= s["pct"] < 80]
        for s in at_risk[:3]:
            insights.append(
                f"🔮 Predicted risk: {s['name']} is at {s['pct']}% — "
                f"may fall below 75% next week."
            )

        # Store as notifications
        for msg in insights:
            existing = conn.execute("""
                SELECT 1 FROM notifications
                WHERE type='ai_insight' AND message=?
                AND date(created_at)=date('now','localtime')
            """, (msg,)).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO notifications(type,message) VALUES(?,?)",
                    ("ai_insight", msg)
                )

        conn.commit()
        conn.close()
        logger.info(f"[SCHEDULER] {len(insights)} AI insights stored.")
    except Exception as e:
        logger.error(f"[SCHEDULER] AI insights failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# Scheduler init
# ═══════════════════════════════════════════════════════════════════════════════
def start_scheduler(app_context=None):
    """Start APScheduler with all jobs. Call from app.py after init_db()."""
    global _scheduler
    if _scheduler and _scheduler.running:
        return _scheduler

    _scheduler = BackgroundScheduler(timezone="Asia/Kolkata",
                                     job_defaults={"misfire_grace_time": 300})

    h2, m2 = map(int, DEFAULTER_CHECK_TIME.split(":"))
    h3, m3 = map(int, DAILY_REPORT_TIME.split(":"))
    hb, mb = map(int, BACKUP_TIME.split(":"))

    # Feature 2 + 10: 6 PM daily
    _scheduler.add_job(job_defaulter_check, CronTrigger(hour=h2, minute=m2),
                       id="defaulter_check", replace_existing=True)

    # Feature 3: 7 PM daily
    _scheduler.add_job(job_daily_email_report, CronTrigger(hour=h3, minute=m3),
                       id="daily_email", replace_existing=True)

    # Feature 5: 11 PM daily
    _scheduler.add_job(job_nightly_backup, CronTrigger(hour=hb, minute=mb),
                       id="nightly_backup", replace_existing=True)

    # Feature 7: 1st of each month at 8 AM
    _scheduler.add_job(job_monthly_report,
                       CronTrigger(day=MONTHLY_REPORT_DAY, hour=8, minute=0),
                       id="monthly_report", replace_existing=True)

    _scheduler.start()
    logger.info("[SCHEDULER] Started — jobs: defaulter@%s, email@%s, backup@%s, monthly@1st",
                DEFAULTER_CHECK_TIME, DAILY_REPORT_TIME, BACKUP_TIME)
    return _scheduler


def stop_scheduler():
    global _scheduler
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)
        logger.info("[SCHEDULER] Stopped.")
